import os
import threading
from datetime import datetime

import pandas as pd
import numpy as np
import customtkinter as ctk
from tkinter import filedialog, messagebox

try:
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# =========================================================================
# FUNÇÕES AUXILIARES
# =========================================================================

def _limpar_nome_coluna(col):
    """
    Limpa nomes de colunas vindos do Excel.
    Remove espaços normais, espaços invisíveis e NBSP.
    """
    return str(col).replace('\xa0', ' ').strip()


def _normalizar_colunas_excel(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza os nomes das colunas sem destruir o DataFrame.
    Ajuda quando vem 'Ação final ', ' MOTIVO ', etc.
    """
    df = df.copy()
    df.columns = [_limpar_nome_coluna(c) for c in df.columns]
    return df


def _achar_coluna(df: pd.DataFrame, opcoes: list, nome_arquivo: str):
    """
    Acha uma coluna aceitando variações de maiúscula/minúscula.
    Exemplo: 'Ação final', 'Ação Final', 'AÇÃO FINAL'.
    """
    mapa = {_limpar_nome_coluna(c).upper(): c for c in df.columns}

    for opcao in opcoes:
        chave = _limpar_nome_coluna(opcao).upper()
        if chave in mapa:
            return mapa[chave]

    raise ValueError(
        f"O arquivo '{nome_arquivo}' não possui nenhuma das colunas esperadas: {opcoes}"
    )


def _validar_colunas(df: pd.DataFrame, obrigatorias: list, nome_arquivo: str):
    faltando = [c for c in obrigatorias if c not in df.columns]
    if faltando:
        raise ValueError(
            f"O arquivo '{nome_arquivo}' está sem as colunas obrigatórias: {faltando}"
        )


def _aplicar_cores(caminho_arquivo: str, col_acao_nome: str):
    """Aplica cores condicionais na coluna de Ação para facilitar leitura visual."""
    if not OPENPYXL_OK:
        return

    import openpyxl

    wb = openpyxl.load_workbook(caminho_arquivo)
    ws = wb.active

    header = [cell.value for cell in ws[1]]

    if col_acao_nome not in header:
        wb.close()
        return

    col_idx = header.index(col_acao_nome) + 1
    col_letra = get_column_letter(col_idx)

    fill_bloqueio = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    fill_alerta = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    fill_manter = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        valor = ws[f"{col_letra}{row}"].value or ""

        if "BLOQUEAR" in str(valor).upper():
            ws[f"{col_letra}{row}"].fill = fill_bloqueio
        elif "ALERTA" in str(valor).upper():
            ws[f"{col_letra}{row}"].fill = fill_alerta
        elif str(valor).strip().upper() == "MANTER":
            ws[f"{col_letra}{row}"].fill = fill_manter

    ws.freeze_panes = "A2"
    wb.save(caminho_arquivo)
    wb.close()


def _data_corte_mes_atual() -> datetime:
    data_atual = datetime.now()
    return datetime(data_atual.year, data_atual.month, 1)


# =========================================================================
# ETAPA 2 - CLASSIFICAÇÃO DE CONTAS
# =========================================================================

def processar_etapa2_core(path_contas: str, path_transferencias: str) -> str:

    # ================= COLUNAS =================
    COL_ID = 'IGA ID'
    COL_CATEGORIA = 'Categoria'
    COL_STATUS = 'Status IGA'
    COL_CRIACAO = 'WhenCreated'
    COL_LOGON = 'LastLogonTimeStamp'

    COL_TRANSFERIDO = 'Transferido?'
    COL_DATA_TRANSF = 'Data da transferência'
    COL_ACAO = 'Ação'
    COL_MOTIVO = 'Motivo Ação'
    COL_OBS = 'Observações Adicionais'

    COL_TRANSF_DATE = 'Transfer Date'
    COL_TRANSF_ID = 'Network ID'

    data_str = datetime.now().strftime('%Y-%m-%d')
    ARQ_SAIDA = f'Relatorio_Contas_FINAL_{data_str}.xlsx'

    # ================= CATEGORIAS =================
    EXCECOES = {'Chave de Processo', 'Conta PAM'}
    DXC = {'Conta A- DXC', 'Conta D- DXC'}
    OUTROS = {'EMPREGADO', 'CONTRATADO', 'Chave A- VALE ou C0', 'Usuário Não Identificado'}

    # ================= CARGA =================
    df = pd.read_excel(path_contas)
    df_transf = pd.read_excel(path_transferencias)

    df = _normalizar_colunas_excel(df)
    df_transf = _normalizar_colunas_excel(df_transf)

    _validar_colunas(
        df,
        [COL_ID, COL_CATEGORIA, COL_STATUS, COL_CRIACAO, COL_LOGON],
        os.path.basename(path_contas)
    )

    linhas_originais = len(df)

    # ================= SAÍDA PADRÃO =================
    for col, padrao in [
        (COL_TRANSFERIDO, 'Não'),
        (COL_DATA_TRANSF, pd.NaT),
        (COL_ACAO, 'Manter'),
        (COL_MOTIVO, ''),
        (COL_OBS, ''),
    ]:
        if col not in df.columns:
            df[col] = padrao

    # ================= DATAS =================
    df[COL_CRIACAO] = pd.to_datetime(df[COL_CRIACAO], errors='coerce')
    df[COL_LOGON] = pd.to_datetime(df[COL_LOGON], errors='coerce')

    # ================= MAPA DE TRANSFERÊNCIAS =================
    if COL_TRANSF_ID in df_transf.columns and COL_ID not in df_transf.columns:
        df_transf = df_transf.rename(columns={COL_TRANSF_ID: COL_ID})

    _validar_colunas(
        df_transf,
        [COL_ID, COL_TRANSF_DATE],
        os.path.basename(path_transferencias)
    )

    df_transf['DataTransferencia'] = pd.to_datetime(df_transf[COL_TRANSF_DATE], errors='coerce')

    mapa_transf = (
        df_transf
        .dropna(subset=[COL_ID])
        .sort_values('DataTransferencia')
        .groupby(COL_ID)['DataTransferencia']
        .last()
    )

    if mapa_transf.empty:
        df[COL_DATA_TRANSF] = pd.NaT
    else:
        df[COL_DATA_TRANSF] = df[COL_ID].map(mapa_transf)
    df[COL_TRANSFERIDO] = df[COL_DATA_TRANSF].notna().map({True: 'Sim', False: 'Não'})

    # ================= MÁSCARAS BASE =================
    mask_excecao = df[COL_CATEGORIA].isin(EXCECOES)
    mask_dxc = df[COL_CATEGORIA].isin(DXC)
    mask_outros = df[COL_CATEGORIA].isin(OUTROS)
    mask_avaliar = ~mask_excecao

    data_corte = _data_corte_mes_atual()
    data_corte_formatada = data_corte.strftime('%d/%m/%Y')

    # ================= CONDIÇÕES =================
    cond_iga = mask_avaliar & (df[COL_STATUS] == 'INACTIVE')

    cond_transfer_base = (
        mask_avaliar &
        df[COL_TRANSFERIDO].eq('Sim') &
        df[COL_DATA_TRANSF].notna() &
        df[COL_CRIACAO].notna()
    )

    dias_ct = (df[COL_DATA_TRANSF] - df[COL_CRIACAO]).dt.days

    cond_bloq_transf = (
        cond_transfer_base &
        (df[COL_CRIACAO] <= df[COL_DATA_TRANSF]) &
        (dias_ct > 7)
    )

    cond_alerta_transf = (
        cond_transfer_base &
        (df[COL_CRIACAO] <= df[COL_DATA_TRANSF]) &
        (dias_ct >= 0) &
        (dias_ct <= 7)
    )

    data_base = df[COL_LOGON].combine_first(df[COL_CRIACAO])
    dias = (data_corte - data_base).dt.days

    cond_dxc_tempo = mask_avaliar & mask_dxc & data_base.notna() & (dias > 180)
    cond_outros_tempo = mask_avaliar & mask_outros & data_base.notna() & (dias > 90)
    cond_tempo = cond_dxc_tempo | cond_outros_tempo

    # ================= PRIORIDADE FINAL =================
    df[COL_ACAO] = 'Manter'
    df[COL_MOTIVO] = ''

    # 1) Alerta de transferência
    df.loc[cond_alerta_transf, COL_ACAO] = 'ALERTA: Verificar Transferência'
    df.loc[cond_alerta_transf, COL_MOTIVO] = 'Criação próxima à data de transferência (possível delay)'

    # 2) Tempo
    df.loc[cond_dxc_tempo, COL_ACAO] = 'BLOQUEAR por TEMPO (DXC > 180 dias)'
    df.loc[cond_dxc_tempo & df[COL_LOGON].isna(), COL_MOTIVO] = (
        'Nunca logado. Conta criada há ' + dias.astype(str) + ' dias'
    )
    df.loc[cond_dxc_tempo & df[COL_LOGON].notna(), COL_MOTIVO] = (
        dias.astype(str) + ' dias sem logon'
    )

    df.loc[cond_outros_tempo, COL_ACAO] = 'BLOQUEAR por TEMPO (Outros > 90 dias)'
    df.loc[cond_outros_tempo & df[COL_LOGON].isna(), COL_MOTIVO] = (
        'Nunca logado. Conta criada há ' + dias.astype(str) + ' dias'
    )
    df.loc[cond_outros_tempo & df[COL_LOGON].notna(), COL_MOTIVO] = (
        dias.astype(str) + ' dias sem logon'
    )

    # 3) Inatividade IGA
    df.loc[cond_iga, COL_ACAO] = 'BLOQUEAR por Inatividade'
    df.loc[cond_iga, COL_MOTIVO] = 'Status IGA = INACTIVE'

    # 4) Transferência com prioridade máxima
    df.loc[cond_bloq_transf, COL_ACAO] = 'BLOQUEAR por Transferência'
    df.loc[cond_bloq_transf, COL_MOTIVO] = 'Conta criada antes da transferência'

    # ================= OBSERVAÇÕES ADICIONAIS =================
    obs = pd.Series([''] * len(df), index=df.index)

    def _add_obs(mask, texto):
        nonlocal obs
        obs = obs.where(~mask, (obs + np.where(obs.eq(''), '', ' | ') + texto))

    e_tempo_final = df[COL_ACAO].isin([
        'BLOQUEAR por TEMPO (DXC > 180 dias)',
        'BLOQUEAR por TEMPO (Outros > 90 dias)'
    ])

    e_iga_final = df[COL_ACAO] == 'BLOQUEAR por Inatividade'
    e_bloq_transf_final = df[COL_ACAO] == 'BLOQUEAR por Transferência'
    e_alerta_final = df[COL_ACAO] == 'ALERTA: Verificar Transferência'

    _add_obs(
        cond_alerta_transf & e_tempo_final,
        'Havia ALERTA de transferência, mas a conta será bloqueada por TEMPO'
    )

    _add_obs(
        cond_alerta_transf & e_iga_final,
        'Havia ALERTA de transferência, mas a conta será bloqueada por INATIVIDADE'
    )

    _add_obs(
        cond_bloq_transf & ~e_bloq_transf_final,
        'Também se enquadra em BLOQUEIO por Transferência'
    )

    _add_obs(
        cond_tempo & ~e_tempo_final & ~e_bloq_transf_final,
        'Também se enquadra em BLOQUEIO por TEMPO'
    )

    _add_obs(
        cond_iga & ~e_iga_final & ~e_bloq_transf_final,
        'Também está INACTIVE no IGA'
    )

    _add_obs(
        cond_alerta_transf & ~e_alerta_final & ~e_tempo_final & ~e_iga_final & ~e_bloq_transf_final,
        'Também se enquadra em ALERTA de transferência'
    )

    df[COL_OBS] = obs

    assert len(df) == linhas_originais, 'ERRO: quantidade de linhas alterada'

    caminho = os.path.join(os.path.dirname(path_contas), ARQ_SAIDA)
    df.to_excel(caminho, index=False)

    _aplicar_cores(caminho, COL_ACAO)

    resumo = df[COL_ACAO].value_counts(dropna=False)
    qtd_sinalizadas = int((df[COL_OBS] != '').sum())

    return (
        f'✅ PLANILHA GERADA COM SUCESSO!\n\n'
        f'📄 Planilha gerada: {ARQ_SAIDA}\n'
        f'📁 Local: {caminho}\n\n'
        f'📅 Data de corte utilizada: {data_corte_formatada}\n\n'
        f'Resumo das Ações:\n{resumo.to_string()}\n\n'
        f'⚠️ Linhas com observação adicional: {qtd_sinalizadas}'
    )


# =========================================================================
# ETAPA 3 - COMPARAÇÃO COM O MÊS ANTERIOR
# =========================================================================

def comparar_etapa3_core(path_atual: str, path_anterior: str) -> str:

    # ================= COLUNAS DE CHAVE =================
    COL_ID = 'IGA ID'
    COL_DOMAIN = 'Domain'
    COL_CN = 'CN'

    # O arquivo pode ter Samaccountname duplicado.
    # O pandas normalmente renomeia a segunda ocorrência como Samaccountname.1.
    POSSIVEIS_SAM = ['Samaccountname', 'SamAccountName', 'SAMAccountName', 'Samaccountname.1']

    # ================= COLUNAS DE AÇÃO =================
    POSSIVEIS_ACAO_FINAL = ['Ação final', 'Ação Final', 'AÇÃO FINAL']
    POSSIVEIS_ACAO_ATUAL = ['Ação final', 'Ação Final', 'AÇÃO FINAL', 'Ação', 'AÇÃO']

    COL_ULTIMA_ACAO = 'Ação final do mês anterior'
    COL_MUDANCA = 'Status da Mudança'

    data_str = datetime.now().strftime('%Y-%m-%d')
    ARQ_SAIDA = f'Relatorio_Comparativo_Etapa3_{data_str}.xlsx'

    df_atual = pd.read_excel(path_atual)
    df_anterior = pd.read_excel(path_anterior)

    df_atual = _normalizar_colunas_excel(df_atual)
    df_anterior = _normalizar_colunas_excel(df_anterior)

    nome_atual = os.path.basename(path_atual)
    nome_anterior = os.path.basename(path_anterior)

    # ================= DESCOBRE COLUNAS REAIS =================
    col_sam_atual = _achar_coluna(df_atual, POSSIVEIS_SAM, nome_atual)
    col_sam_anterior = _achar_coluna(df_anterior, POSSIVEIS_SAM, nome_anterior)

    col_acao_atual = _achar_coluna(df_atual, POSSIVEIS_ACAO_ATUAL, nome_atual)

    # Mês anterior tem que usar Ação final.
    # Não usa Ação comum para não puxar decisão errada.
    col_acao_anterior = _achar_coluna(df_anterior, POSSIVEIS_ACAO_FINAL, nome_anterior)

    # ================= VALIDA COLUNAS =================
    _validar_colunas(
        df_atual,
        [COL_ID, COL_DOMAIN, COL_CN, col_sam_atual, col_acao_atual],
        nome_atual
    )

    _validar_colunas(
        df_anterior,
        [COL_ID, COL_DOMAIN, COL_CN, col_sam_anterior, col_acao_anterior],
        nome_anterior
    )

    linhas_originais = len(df_atual)

    # ================= NORMALIZA VALORES =================
    def normalizar_valor(serie):
        return (
            serie
            .fillna('')
            .astype(str)
            .str.replace('\xa0', ' ', regex=False)
            .str.strip()
            .str.upper()
        )

    # ================= CHAVE COMPOSTA =================
    # Não usa só IGA ID.
    # Usa pelo menos 4 campos para identificar a linha/conta:
    # IGA ID + Domain + Samaccountname + CN
    def criar_chave(df, col_sam):
        return (
            normalizar_valor(df[COL_ID])
            + '||'
            + normalizar_valor(df[COL_DOMAIN])
            + '||'
            + normalizar_valor(df[col_sam])
            + '||'
            + normalizar_valor(df[COL_CN])
        )

    df_atual['_CHAVE_COMPARACAO'] = criar_chave(df_atual, col_sam_atual)
    df_anterior['_CHAVE_COMPARACAO'] = criar_chave(df_anterior, col_sam_anterior)

    # ================= FILTRA LINHAS VÁLIDAS DO MÊS ANTERIOR =================
    df_anterior_validas = df_anterior[
        df_anterior[COL_ID].notna()
        & df_anterior[COL_DOMAIN].notna()
        & df_anterior[col_sam_anterior].notna()
        & df_anterior[COL_CN].notna()
    ].copy()

    # Se mesmo com a chave composta tiver duplicidade, mantém a última ocorrência.
    # Isso evita estourar linha no merge e garante que a quantidade de linhas do atual não muda.
    mapa_acao_anterior = (
        df_anterior_validas
        .drop_duplicates(subset=['_CHAVE_COMPARACAO'], keep='last')
        .set_index('_CHAVE_COMPARACAO')[col_acao_anterior]
    )

    # ================= TRAZ SOMENTE AÇÃO FINAL DO MÊS ANTERIOR =================
    df_atual[COL_ULTIMA_ACAO] = df_atual['_CHAVE_COMPARACAO'].map(mapa_acao_anterior)

    estava_no_mes_anterior = df_atual[COL_ULTIMA_ACAO].notna()

    df_atual.loc[~estava_no_mes_anterior, COL_ULTIMA_ACAO] = 'Não estava no mês anterior'

    # ================= COMPARA ATUAL X MÊS ANTERIOR =================
    acao_atual_norm = (
        df_atual[col_acao_atual]
        .fillna('')
        .astype(str)
        .str.replace('\xa0', ' ', regex=False)
        .str.strip()
    )

    acao_anterior_norm = (
        df_atual[COL_ULTIMA_ACAO]
        .fillna('')
        .astype(str)
        .str.replace('\xa0', ' ', regex=False)
        .str.strip()
    )

    mudou = estava_no_mes_anterior & (acao_atual_norm != acao_anterior_norm)
    igual = estava_no_mes_anterior & (acao_atual_norm == acao_anterior_norm)

    df_atual[COL_MUDANCA] = 'Nova conta no mês atual'
    df_atual.loc[igual, COL_MUDANCA] = 'Sem alteração'
    df_atual.loc[mudou, COL_MUDANCA] = (
        df_atual.loc[mudou, COL_ULTIMA_ACAO].astype(str)
        + ' → '
        + df_atual.loc[mudou, col_acao_atual].astype(str)
    )

    # Remove auxiliar
    df_atual = df_atual.drop(columns=['_CHAVE_COMPARACAO'])

    assert len(df_atual) == linhas_originais, 'ERRO: quantidade de linhas alterada'

    caminho = os.path.join(os.path.dirname(path_atual), ARQ_SAIDA)
    df_atual.to_excel(caminho, index=False)

    _aplicar_cores(caminho, col_acao_atual)

    novas = int((~estava_no_mes_anterior).sum())
    mudaram = int(mudou.sum())
    sem_alteracao = int(igual.sum())

    return (
        f'✅ PLANILHA GERADA COM SUCESSO!\n\n'
        f'📄 Planilha gerada: {ARQ_SAIDA}\n'
        f'📁 Local: {caminho}\n\n'
        f'Coluna usada no mês atual: {col_acao_atual}\n'
        f'Coluna usada do mês anterior: {col_acao_anterior}\n'
        f'Chave de comparação: IGA ID + Domain + {col_sam_atual} + CN\n\n'
        f'Total de linhas: {linhas_originais}\n'
        f'Novas contas no mês atual: {novas}\n'
        f'Mudaram de ação: {mudaram}\n'
        f'Sem alteração: {sem_alteracao}'
    )


# =========================================================================
# GUI
# =========================================================================

APP_TITLE = "Processador de Contas - Etapa 2 e Etapa 3"
DEFAULT_GEOMETRY = "800x560"


class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("System")
        ctk.set_default_color_theme("blue")

        self.title(APP_TITLE)
        self.geometry(DEFAULT_GEOMETRY)
        self.minsize(760, 520)

        self.tabview = ctk.CTkTabview(self, corner_radius=16)
        self.tabview.pack(fill="both", expand=True, padx=20, pady=20)

        self.tab_etapa2 = self.tabview.add("Etapa 2 - Classificação")
        self.tab_etapa3 = self.tabview.add("Etapa 3 - Comparativo")

        self._build_etapa2(self.tab_etapa2)
        self._build_etapa3(self.tab_etapa3)

    # ---------------- ETAPA 2 ----------------
    def _build_etapa2(self, parent):
        self.path_contas = ctk.StringVar(value="")
        self.path_transferencias = ctk.StringVar(value="")
        self.planilha_gerada2 = ctk.StringVar(value="Nenhuma planilha gerada ainda.")

        title = ctk.CTkLabel(
            parent,
            text="Processar Relatório (Etapa 2)",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title.pack(anchor="w", padx=6, pady=(10, 6))

        subtitle = ctk.CTkLabel(
            parent,
            text=(
                "Escolha os arquivos .xlsx de Contas e Transferências. "
                "Os dias sem logon são calculados a partir do primeiro dia do mês atual."
            ),
            font=ctk.CTkFont(size=13),
            wraplength=720,
            justify="left"
        )
        subtitle.pack(anchor="w", padx=6, pady=(0, 14))

        self._file_picker(
            parent,
            "Arquivo de Contas (.xlsx)",
            self.path_contas,
            lambda: self._pick_file(self.path_contas, "Contas")
        )

        self._file_picker(
            parent,
            "Arquivo de Transferências (.xlsx)",
            self.path_transferencias,
            lambda: self._pick_file(self.path_transferencias, "Transferências")
        )

        actions = ctk.CTkFrame(parent, fg_color="transparent")
        actions.pack(fill="x", padx=6, pady=(14, 8))

        self.btn_processar = ctk.CTkButton(
            actions,
            text="Processar",
            height=42,
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._on_processar_etapa2
        )
        self.btn_processar.pack(side="left")

        self.progress2 = ctk.CTkProgressBar(actions, mode="indeterminate", height=12)
        self.progress2.pack(side="left", fill="x", expand=True, padx=(14, 0))
        self.progress2.stop()

        self.lbl_planilha2 = ctk.CTkLabel(
            parent,
            textvariable=self.planilha_gerada2,
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#2E7D32"
        )
        self.lbl_planilha2.pack(anchor="w", padx=10, pady=(0, 6))

        self.status_box2 = self._status_box(parent)
        self._set_status(self.status_box2, "Selecione os dois arquivos e clique em Processar.")

    # ---------------- ETAPA 3 ----------------
    def _build_etapa3(self, parent):
        self.path_atual = ctk.StringVar(value="")
        self.path_anterior = ctk.StringVar(value="")
        self.planilha_gerada3 = ctk.StringVar(value="Nenhuma planilha gerada ainda.")

        title = ctk.CTkLabel(
            parent,
            text="Comparar com o mês anterior (Etapa 3)",
            font=ctk.CTkFont(size=20, weight="bold")
        )
        title.pack(anchor="w", padx=6, pady=(10, 6))

        subtitle = ctk.CTkLabel(
            parent,
            text=(
                "Escolha o relatório do mês ATUAL e o do mês ANTERIOR. "
                "A comparação será feita por IGA ID + Domain + Samaccountname + CN, "
                "trazendo somente a coluna Ação final do mês anterior."
            ),
            font=ctk.CTkFont(size=13),
            wraplength=720,
            justify="left"
        )
        subtitle.pack(anchor="w", padx=6, pady=(0, 14))

        self._file_picker(
            parent,
            "Relatório do mês ATUAL (.xlsx)",
            self.path_atual,
            lambda: self._pick_file(self.path_atual, "Mês Atual")
        )

        self._file_picker(
            parent,
            "Relatório do mês ANTERIOR (.xlsx)",
            self.path_anterior,
            lambda: self._pick_file(self.path_anterior, "Mês Anterior")
        )

        actions = ctk.CTkFrame(parent, fg_color="transparent")
        actions.pack(fill="x", padx=6, pady=(14, 8))

        self.btn_comparar = ctk.CTkButton(
            actions,
            text="Comparar",
            height=42,
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._on_comparar_etapa3
        )
        self.btn_comparar.pack(side="left")

        self.progress3 = ctk.CTkProgressBar(actions, mode="indeterminate", height=12)
        self.progress3.pack(side="left", fill="x", expand=True, padx=(14, 0))
        self.progress3.stop()

        self.lbl_planilha3 = ctk.CTkLabel(
            parent,
            textvariable=self.planilha_gerada3,
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color="#2E7D32"
        )
        self.lbl_planilha3.pack(anchor="w", padx=10, pady=(0, 6))

        self.status_box3 = self._status_box(parent)
        self._set_status(self.status_box3, "Selecione os dois arquivos e clique em Comparar.")

    # ---------------- HELPERS DE UI ----------------
    def _file_picker(self, parent, label, var, command):
        row = ctk.CTkFrame(parent, corner_radius=12)
        row.pack(fill="x", padx=6, pady=6)

        lbl = ctk.CTkLabel(
            row,
            text=label,
            font=ctk.CTkFont(size=13, weight="bold")
        )
        lbl.pack(anchor="w", padx=12, pady=(10, 4))

        inner = ctk.CTkFrame(row, fg_color="transparent")
        inner.pack(fill="x", padx=12, pady=(0, 12))

        entry = ctk.CTkEntry(
            inner,
            textvariable=var,
            placeholder_text="Selecione um arquivo..."
        )
        entry.pack(side="left", fill="x", expand=True)

        btn = ctk.CTkButton(
            inner,
            text="Selecionar",
            width=120,
            command=command
        )
        btn.pack(side="left", padx=(10, 0))

    def _status_box(self, parent):
        status_frame = ctk.CTkFrame(parent, corner_radius=12)
        status_frame.pack(fill="both", expand=True, padx=6, pady=(10, 6))

        status_label = ctk.CTkLabel(
            status_frame,
            text="Status",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        status_label.pack(anchor="w", padx=14, pady=(12, 6))

        box = ctk.CTkTextbox(status_frame, corner_radius=10, wrap="word")
        box.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        return box

    def _set_status(self, box, text):
        box.configure(state="normal")
        box.delete("1.0", "end")
        box.insert("1.0", text)
        box.configure(state="disabled")

    def _pick_file(self, var, label):
        path = filedialog.askopenfilename(
            title=f"Selecione o arquivo de {label}",
            filetypes=[("Excel", "*.xlsx *.xls")]
        )

        if path:
            var.set(path)

    # ---------------- AÇÕES ETAPA 2 ----------------
    def _on_processar_etapa2(self):
        pc = self.path_contas.get().strip()
        pt = self.path_transferencias.get().strip()

        if not pc or not os.path.exists(pc):
            self._set_status(self.status_box2, "❌ Selecione um arquivo válido de Contas.")
            return

        if not pt or not os.path.exists(pt):
            self._set_status(self.status_box2, "❌ Selecione um arquivo válido de Transferências.")
            return

        self.planilha_gerada2.set("Processando...")
        self.btn_processar.configure(state="disabled", text="Processando...")
        self.progress2.start()
        self._set_status(self.status_box2, "🔄 Processando...")

        threading.Thread(
            target=self._run_etapa2,
            args=(pc, pt),
            daemon=True
        ).start()

    def _run_etapa2(self, pc, pt):
        try:
            resultado = processar_etapa2_core(pc, pt)
        except Exception as e:
            resultado = f"❌ ERRO inesperado:\n{e}"

        self.after(0, lambda: self._finish_etapa2(resultado))

    def _finish_etapa2(self, resultado):
        self.progress2.stop()
        self.btn_processar.configure(state="normal", text="Processar")
        self._set_status(self.status_box2, resultado)

        if resultado.startswith("✅"):
            self.planilha_gerada2.set("✅ PLANILHA DA ETAPA 2 GERADA COM SUCESSO")
            messagebox.showinfo("Sucesso", "Planilha da Etapa 2 gerada com sucesso!")
        else:
            self.planilha_gerada2.set("❌ Erro ao gerar planilha da Etapa 2")

    # ---------------- AÇÕES ETAPA 3 ----------------
    def _on_comparar_etapa3(self):
        pa = self.path_atual.get().strip()
        pant = self.path_anterior.get().strip()

        if not pa or not os.path.exists(pa):
            self._set_status(self.status_box3, "❌ Selecione um arquivo válido do mês Atual.")
            return

        if not pant or not os.path.exists(pant):
            self._set_status(self.status_box3, "❌ Selecione um arquivo válido do mês Anterior.")
            return

        self.planilha_gerada3.set("Comparando...")
        self.btn_comparar.configure(state="disabled", text="Comparando...")
        self.progress3.start()
        self._set_status(self.status_box3, "🔄 Comparando...")

        threading.Thread(
            target=self._run_etapa3,
            args=(pa, pant),
            daemon=True
        ).start()

    def _run_etapa3(self, pa, pant):
        try:
            resultado = comparar_etapa3_core(pa, pant)
        except Exception as e:
            resultado = f"❌ ERRO inesperado:\n{e}"

        self.after(0, lambda: self._finish_etapa3(resultado))

    def _finish_etapa3(self, resultado):
        self.progress3.stop()
        self.btn_comparar.configure(state="normal", text="Comparar")
        self._set_status(self.status_box3, resultado)

        if resultado.startswith("✅"):
            self.planilha_gerada3.set("✅ PLANILHA DA ETAPA 3 GERADA COM SUCESSO")
            messagebox.showinfo("Sucesso", "Planilha da Etapa 3 gerada com sucesso!")
        else:
            self.planilha_gerada3.set("❌ Erro ao gerar planilha da Etapa 3")


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
