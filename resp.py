import re
import threading
import tkinter as tk
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from tkinter import filedialog, messagebox

import customtkinter as ctk
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# ── Tema ──────────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

PRIMARY   = "#00c5b6"
PRIMARY_D = "#009e91"
BG        = "#f4f5f7"
CARD_BG   = "#ffffff"
BORDER    = "#d8dce4"
TEXT_H    = "#1a1f2e"
TEXT_MUTE = "#7a8099"
BADGE_BG  = "#e0faf8"
BADGE_FG  = "#00857a"
STAT_FG   = "#00c5b6"
INFO_BG   = "#e8f7f6"
INFO_FG   = "#00796b"
SUCC_BG   = "#e8f5e9"
SUCC_FG   = "#2e7d32"
ERR_BG    = "#fdecea"
ERR_FG    = "#c0392b"
WARN_BG   = "#fff8e1"

TODAY          = datetime.now(timezone.utc)
LIMITE_CREATED = pd.Timestamp("2026-06-30 23:59:59", tz="UTC")
ACAO_CHOICES   = ["Bloquear", "Manter"]

# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean(val) -> str:
    s = str(val).strip() if pd.notna(val) else ""
    return "" if s.lower() in ("nan", "none", "nat", "") else s

def _norm(name: str) -> str:
    return str(name).strip().lower().replace("/", " ").replace("-", " ").replace("_", " ")

def _get_col(row, nomes, default=""):
    if isinstance(nomes, str):
        nomes = [nomes]
    for n in nomes:
        if n in row.index:
            return row.get(n, default)
    normalized = {_norm(c): c for c in row.index}
    for n in nomes:
        col = normalized.get(_norm(n))
        if col:
            return row.get(col, default)
    return default

def _parse_data(val):
    if pd.isna(val):
        return pd.NaT
    s = str(val).strip()
    if s.lower() in ("", "nan", "none", "nat", "não encontrado", "nao encontrado"):
        return pd.NaT
    return pd.to_datetime(s, errors="coerce", dayfirst=True, utc=True)

def _is_nunca(val) -> bool:
    return str(val).strip().upper() == "NUNCA LOGOU"

def _dias(val):
    s = str(val).strip().upper()
    if s in ("NUNCA LOGOU", "", "NAN", "NONE", "NAT"):
        return None
    try:
        return int(float(s))
    except Exception:
        return None

def _menor_dias(*dias):
    vals = [v for v in dias if v is not None]
    return min(vals) if vals else None

def _classificar(row) -> str:
    workflow = _clean(_get_col(row, ["Functional Request Workflow"], ""))
    if workflow.strip().lower() != "completed":
        return "NÃO ELEGÍVEL"
    created = _parse_data(_get_col(row, ["Created", "Data de criação", "whenCreated"], ""))
    if pd.notna(created) and created > LIMITE_CREATED:
        return "NÃO ELEGÍVEL"
    ad_ll  = _is_nunca(_get_col(row, ["Dias sem logar - lastLogon - AD"], ""))
    ad_lts = _is_nunca(_get_col(row, ["Dias sem Logar - lastLogonTimestamp - AD",
                                       "Dias sem logar - lastLogonTimestamp - AD"], ""))
    az     = _is_nunca(_get_col(row, ["Dias sem logar - Azure"], ""))
    if ad_ll and ad_lts and az:
        return "NUNCA LOGOU"
    menor = _menor_dias(
        _dias(_get_col(row, ["Dias sem logar - lastLogon - AD"], "")),
        _dias(_get_col(row, ["Dias sem Logar - lastLogonTimestamp - AD",
                              "Dias sem logar - lastLogonTimestamp - AD"], "")),
        _dias(_get_col(row, ["Dias sem logar - Azure"], "")),
    )
    if menor is not None and menor > 365:
        return "SEM LOGIN +1ANO"
    return "NÃO ELEGÍVEL"

def _responsavel_ativo(row) -> tuple[str, str]:
    slots = [
        ("Responsável 1", ["UAC Responsável 1", "UAC/STATUS Responsável 1"]),
        ("Responsável 2", ["UAC Responsável 2", "UAC/STATUS Responsável 2"]),
        ("Responsável 3", ["UAC Responsável 3", "UAC/STATUS Responsável 3"]),
    ]
    for email_col, uac_cols in slots:
        email = _clean(_get_col(row, [email_col], "")).lower()
        uac   = _clean(_get_col(row, uac_cols, ""))
        if not email:
            continue
        try:
            uac_int = int(float(uac)) if uac else None
        except Exception:
            uac_int = None
        if uac_int == 512:
            return email, uac
    return "", ""

def _sam(row) -> str:
    return _clean(_get_col(row, ["SamAccountName", "Conta", "Conta (SamAccountName)"], ""))

def _created_fmt(row) -> str:
    raw = _get_col(row, ["whenCreated", "Created", "Data de criação"], "")
    dt  = _parse_data(raw)
    return dt.strftime("%d/%m/%Y") if pd.notna(dt) else _clean(raw)

def _dominio(row) -> str:
    return _clean(_get_col(row, ["Domínio", "Dominio", "Domain"], ""))

def _request_id(row) -> str:
    return _clean(_get_col(row, ["RequestID", "Request ID", "REQUEST ID", "Request Id"], ""))

def _safe_filename(text: str, max_len: int = 80) -> str:
    text = _clean(text).lower().replace("@", "_at_").replace(".", "_")
    text = re.sub(r"[^a-zA-Z0-9_\-]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text[:max_len] or "sem_responsavel"

def _formatar(ws, df, header_color="00C5B6"):
    hdr_fill = PatternFill("solid", start_color=header_color, end_color=header_color)
    hdr_font = Font(bold=True, color="FFFFFF", name="Segoe UI", size=11)
    zebra    = PatternFill("solid", start_color="E0FAF8", end_color="E0FAF8")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row_idx in range(2, len(df) + 2):
        if row_idx % 2 == 0:
            for c in range(1, len(df.columns) + 1):
                ws.cell(row=row_idx, column=c).fill = zebra
        for c in range(1, len(df.columns) + 1):
            ws.cell(row=row_idx, column=c).alignment = Alignment(vertical="top")
    for idx, col in enumerate(df.columns, start=1):
        max_len = max(len(str(col)), max((len(str(v)) for v in df[col].head(300)), default=0))
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = min(max(max_len + 2, 12), 55)

def _gerar_individual(df_ind: pd.DataFrame) -> bytes:
    """Gera Excel individual por responsável."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_ind.to_excel(writer, index=False, sheet_name="Contas")
        ws = writer.book["Contas"]
        _formatar(ws, df_ind)
        for col, w in {"A": 32, "B": 14, "C": 36, "D": 22, "E": 18,
                       "F": 16, "G": 16, "H": 14, "I": 45}.items():
            ws.column_dimensions[col].width = w
        n = len(df_ind)
        if n > 0:
            dv = DataValidation(
                type="list",
                formula1='"' + ",".join(ACAO_CHOICES) + '"',
                allow_blank=True,
                showDropDown=False,
            )
            ws.add_data_validation(dv)
            dv.sqref = f"H2:H{n + 1}"
    return buf.getvalue()

# ── Lógica principal ──────────────────────────────────────────────────────────

def gerar_planilhas(df_orig: pd.DataFrame) -> tuple[dict[str, bytes], dict]:
    df = df_orig.copy()
    df.columns = df.columns.str.strip()

    obrigatorias = [
        "SamAccountName", "Functional Request Workflow",
        "Dias sem logar - lastLogon - AD", "Dias sem Logar - lastLogonTimestamp - AD",
        "Dias sem logar - Azure", "Responsável 1", "Responsável 2", "Responsável 3",
    ]
    faltantes = [c for c in obrigatorias if c not in df.columns]
    if faltantes:
        raise Exception("Colunas ausentes: " + ", ".join(faltantes))

    df["_classif"]     = df.apply(_classificar, axis=1)
    df[["_resp_email", "_resp_uac"]] = df.apply(
        lambda r: pd.Series(_responsavel_ativo(r)), axis=1
    )

    # Agrupar elegíveis por responsável ativo
    df_el = df[df["_classif"].isin(["NUNCA LOGOU", "SEM LOGIN +1ANO"])].copy()
    grupos: dict[str, list[dict]] = {}
    for _, row in df_el.iterrows():
        resp = row["_resp_email"] or "SEM RESPONSÁVEL ATIVO"
        grupos.setdefault(resp, []).append(row.to_dict())

    # Responsável entra no lote agora só se tiver >= 1 NUNCA LOGOU
    enviar   = {r: c for r, c in grupos.items() if any(x["_classif"] == "NUNCA LOGOU" for x in c)}
    aguardar = {r: c for r, c in grupos.items() if r not in enviar}

    # Nomes de arquivo únicos
    usados: set[str] = set()
    nomes: dict[str, str] = {}
    for resp in sorted(enviar):
        base = "validar_responsavel_ativo" if resp == "SEM RESPONSÁVEL ATIVO" else f"contas_{_safe_filename(resp)}"
        nome = f"{base}.xlsx"
        i = 2
        while nome in usados:
            nome = f"{base}_{i}.xlsx"
            i += 1
        usados.add(nome)
        nomes[resp] = nome

    # Gerar individuais (exceto sem responsável ativo)
    arquivos: dict[str, bytes] = {}
    for resp, contas in enviar.items():
        if resp == "SEM RESPONSÁVEL ATIVO":
            continue

        rows = []
        for c in contas:
            s = pd.Series(c)

            ll_ad_dt = _parse_data(_get_col(s, ["lastLogonTimestamp", "lastLogon"], ""))
            ll_ad    = ll_ad_dt.strftime("%d/%m/%Y") if pd.notna(ll_ad_dt) else "NUNCA LOGOU"

            ll_az_dt = _parse_data(_get_col(s, ["lastLogon - Azure", "LastLogon - Azure"], ""))
            ll_az    = ll_az_dt.strftime("%d/%m/%Y") if pd.notna(ll_az_dt) else "NUNCA LOGOU"

            rows.append({
                "SamAccountName":  _sam(s),
                "whenCreated":     _created_fmt(s),
                "CreatedBy":       _clean(_get_col(s, ["CreatedBy"], "")),
                "Domínio":         _dominio(s),
                "Request ID":      _request_id(s),
                "LastLogon AD":    ll_ad,
                "LastLogon Azure": ll_az,
                "Ação":            "",
                "Motivo":          "",
            })

        df_ind = pd.DataFrame(rows, columns=[
            "SamAccountName", "whenCreated", "CreatedBy", "Domínio",
            "Request ID", "LastLogon AD", "LastLogon Azure", "Ação", "Motivo",
        ])
        arquivos[nomes[resp]] = _gerar_individual(df_ind)

    stats = {
        "nunca":    int((df["_classif"] == "NUNCA LOGOU").sum()),
        "mais1":    int((df["_classif"] == "SEM LOGIN +1ANO").sum()),
        "nao_el":   int((df["_classif"] == "NÃO ELEGÍVEL").sum()),
        "sem_resp": int(
            (df["_classif"].isin(["NUNCA LOGOU", "SEM LOGIN +1ANO"]) & (df["_resp_email"] == "")).sum()
        ),
        "resp_enviar":   len([r for r in enviar   if r != "SEM RESPONSÁVEL ATIVO"]),
        "resp_aguardar": len([r for r in aguardar if r != "SEM RESPONSÁVEL ATIVO"]),
        "arquivos":      len(arquivos),
    }
    return arquivos, stats

# ── UI ────────────────────────────────────────────────────────────────────────

def make_card(parent):
    return ctk.CTkFrame(parent, fg_color=CARD_BG, corner_radius=12,
                        border_width=1, border_color=BORDER)

def primary_btn(parent, text, cmd, width=240):
    return ctk.CTkButton(parent, text=text, command=cmd, fg_color=PRIMARY,
                         hover_color=PRIMARY_D, text_color="#fff", corner_radius=8,
                         font=("Segoe UI", 13, "bold"), width=width, height=38)

def upload_btn(parent, text, cmd, width=220):
    return ctk.CTkButton(parent, text=text, command=cmd, fg_color=INFO_BG,
                         hover_color="#d0efed", text_color=INFO_FG, corner_radius=8,
                         border_width=1, border_color=PRIMARY,
                         font=("Segoe UI", 12), width=width, height=34)


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Planilhas por Responsável · Vale IAM")
        self.geometry("860x620")
        self.minsize(720, 500)
        self.configure(fg_color=BG)
        self._file_path = None
        self._file_var  = tk.StringVar(value="Nenhum arquivo selecionado")
        self._output_dir = None
        self._dir_var   = tk.StringVar(value="Nenhuma pasta selecionada")
        self._stats = None
        self._build()

    def _build(self):
        p = ctk.CTkScrollableFrame(self, fg_color=BG, corner_radius=0)
        p.pack(fill="both", expand=True)
        self._p = p

        ctk.CTkLabel(p, text="Planilhas por Responsável",
                     font=("Segoe UI", 22, "bold"), text_color=TEXT_H
                     ).pack(anchor="w", padx=32, pady=(28, 4))
        ctk.CTkLabel(p,
                     text="Agrupa contas elegíveis por responsável ativo (UAC 512) e gera uma\n"
                          "planilha individual por responsável com Ação e Motivo.",
                     font=("Segoe UI", 12), text_color=TEXT_MUTE, justify="left"
                     ).pack(anchor="w", padx=32, pady=(0, 20))

        # Card planilha
        c1 = make_card(p)
        c1.pack(fill="x", padx=32, pady=4)
        top1 = ctk.CTkFrame(c1, fg_color="transparent")
        top1.pack(fill="x", padx=18, pady=(16, 8))
        ctk.CTkLabel(top1, text="PLANILHA IAM KEY CLEANER",
                     fg_color=BADGE_BG, text_color=BADGE_FG, corner_radius=6,
                     font=("Courier New", 11, "bold"), padx=8, pady=3).pack(side="left")
        ctk.CTkLabel(top1, text="Saída da Etapa 2",
                     font=("Segoe UI", 14, "bold"), text_color=TEXT_H).pack(side="left", padx=12)
        ctk.CTkLabel(c1,
                     text="Colunas necessárias: SamAccountName · whenCreated · CreatedBy · Domínio · Request ID\n"
                          "lastLogonTimestamp · lastLogon · lastLogon - Azure\n"
                          "Dias sem logar - lastLogon - AD · Dias sem Logar - lastLogonTimestamp - AD · Dias sem logar - Azure\n"
                          "Responsável 1/2/3 · UAC Responsável 1/2/3 · Functional Request Workflow",
                     font=("Segoe UI", 11), text_color=TEXT_MUTE, justify="left"
                     ).pack(anchor="w", padx=18, pady=(0, 10))
        upload_btn(c1, "  Selecionar arquivo (.xlsx)", self._pick_file).pack(padx=18, pady=(0, 8))
        ctk.CTkLabel(c1, textvariable=self._file_var, font=("Segoe UI", 10),
                     text_color=TEXT_MUTE, wraplength=700, justify="left"
                     ).pack(anchor="w", padx=18, pady=(0, 14))

        # Card pasta
        c2 = make_card(p)
        c2.pack(fill="x", padx=32, pady=4)
        top2 = ctk.CTkFrame(c2, fg_color="transparent")
        top2.pack(fill="x", padx=18, pady=(16, 8))
        ctk.CTkLabel(top2, text="PASTA DE SAÍDA",
                     fg_color=BADGE_BG, text_color=BADGE_FG, corner_radius=6,
                     font=("Courier New", 11, "bold"), padx=8, pady=3).pack(side="left")
        ctk.CTkLabel(top2, text="Onde salvar as planilhas individuais",
                     font=("Segoe UI", 14, "bold"), text_color=TEXT_H).pack(side="left", padx=12)
        ctk.CTkLabel(c2, text="Uma planilha por responsável será salva nesta pasta.",
                     font=("Segoe UI", 11), text_color=TEXT_MUTE, justify="left"
                     ).pack(anchor="w", padx=18, pady=(0, 10))
        upload_btn(c2, "  Selecionar pasta", self._pick_dir).pack(padx=18, pady=(0, 8))
        ctk.CTkLabel(c2, textvariable=self._dir_var, font=("Segoe UI", 10),
                     text_color=TEXT_MUTE, wraplength=700, justify="left"
                     ).pack(anchor="w", padx=18, pady=(0, 14))

        self._btn_proc = primary_btn(p, "    Gerar planilhas", self._processar, width=240)
        self._btn_proc.pack(pady=18)

        self._sf = ctk.CTkFrame(p, fg_color="transparent")
        self._sf.pack(fill="x", padx=32)

    def _pick_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if path:
            self._file_path = path
            self._file_var.set(path.split("/")[-1])

    def _pick_dir(self):
        path = filedialog.askdirectory()
        if path:
            self._output_dir = path
            self._dir_var.set(path)

    def _processar(self):
        if not self._file_path:
            messagebox.showerror("Arquivo faltando", "Selecione a planilha antes de processar.")
            return
        if not self._output_dir:
            messagebox.showerror("Pasta faltando", "Selecione a pasta de saída.")
            return
        self._btn_proc.configure(state="disabled", text="Processando…")
        for w in self._sf.winfo_children():
            w.destroy()
        prog = ctk.CTkProgressBar(self._sf, mode="indeterminate")
        prog.pack(fill="x", pady=8)
        prog.start()
        self._prog = prog
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        try:
            df = pd.read_excel(self._file_path, engine="openpyxl")
            df.columns = df.columns.str.strip()
            arquivos, stats = gerar_planilhas(df)
            out = Path(self._output_dir)
            for nome, conteudo in arquivos.items():
                (out / nome).write_bytes(conteudo)
            self._stats = stats
            self.after(0, self._done)
        except Exception as e:
            import traceback
            self.after(0, lambda m=str(e) + "\n\n" + traceback.format_exc(): self._error(m))

    def _done(self):
        self._prog.stop()
        self._prog.destroy()
        self._btn_proc.configure(state="normal", text="    Gerar planilhas")
        self._render_resultado()

    def _error(self, msg):
        self._prog.stop()
        self._prog.destroy()
        self._btn_proc.configure(state="normal", text="    Gerar planilhas")
        ctk.CTkLabel(self._sf, text=f"Erro:\n{msg}", fg_color=ERR_BG, text_color=ERR_FG,
                     corner_radius=8, padx=14, pady=10, font=("Courier New", 10),
                     wraplength=800, justify="left").pack(fill="x", pady=4)

    def _render_resultado(self):
        for w in self._sf.winfo_children():
            w.destroy()
        s = self._stats

        ctk.CTkFrame(self._sf, height=1, fg_color=BORDER).pack(fill="x", pady=8)
        ctk.CTkLabel(self._sf, text="Resultado", font=("Segoe UI", 16, "bold"),
                     text_color=TEXT_H).pack(anchor="w", pady=(0, 10))

        sf = ctk.CTkFrame(self._sf, fg_color="transparent")
        sf.pack(fill="x", pady=(0, 14))
        sf.columnconfigure((0, 1, 2, 3, 4, 5, 6), weight=1, uniform="s")

        for i, (num, lbl, bg) in enumerate([
            (s["nunca"],          "Nunca logou",           BADGE_BG),
            (s["mais1"],          "Sem login +1 ano",      INFO_BG),
            (s["nao_el"],         "Não elegível",          "#f5f5f5"),
            (s["sem_resp"],       "Sem resp. ativo",       WARN_BG),
            (s["resp_enviar"],    "Resp. p/ contato",      SUCC_BG),
            (s["resp_aguardar"],  "Resp. aguardar",        WARN_BG),
            (s["arquivos"],       "Planilhas geradas",     SUCC_BG),
        ]):
            box = ctk.CTkFrame(sf, fg_color=bg, corner_radius=8,
                               border_width=1, border_color=BORDER)
            box.grid(row=0, column=i, padx=4, sticky="ew")
            ctk.CTkLabel(box, text=str(num), font=("Courier New", 22, "bold"),
                         text_color=STAT_FG).pack(pady=(10, 2))
            ctk.CTkLabel(box, text=lbl, font=("Segoe UI", 10), text_color=TEXT_MUTE,
                         wraplength=110, justify="center").pack(pady=(0, 10))

        ctk.CTkLabel(
            self._sf,
            text=f"✔  {s['arquivos']} planilha(s) salva(s) em:\n   {self._output_dir}",
            fg_color=SUCC_BG, text_color=SUCC_FG, corner_radius=8,
            font=("Segoe UI", 12), justify="left", padx=16, pady=12, wraplength=780,
        ).pack(fill="x", pady=8)


if __name__ == "__main__":
    app = App()
    app.mainloop()
